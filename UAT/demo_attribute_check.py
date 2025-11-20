from ast import main
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, SessionNotCreatedException, StaleElementReferenceException, TimeoutException,ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver.common.by import By
import time
from SelectCertificate import authenticate_with_certificate
from selenium.webdriver.chrome.service import Service
from datetime import datetime
import win32com.client as win32
import pandas as pd
import openpyxl
import threading
import pygetwindow as gw
import pyautogui
import os
import openpyxl
from pynput.keyboard import Key, Controller
# username='demo_japanese_distributor@pproap.com'venv
# password = "Login2PRP!"
# base_url = "https://partner.hpe.com/web/prp"
# excel_file_path = r'D:\UAT\Demo_account_attribute\validate.xlsx'

# webdriver_path = "Webdrivers\\chromedriver.exe"
# service = Service(webdriver_path)
# driver = webdriver.Chrome(service=service)
# driver.maximize_window()

def setUp():
    global excel_file_path, driver
    webdriver_path="Webdrivers\\chromedriver.exe"
    service = Service(webdriver_path)
    driver=webdriver.Chrome(service=service)
    driver.maximize_window()      

def login(demo_account):
    
    password = "Login2PRP!"
    base_url = "https://partner.hpe.com/web/prp"
    # excel_file_path = r'D:\UAT\Demo_account_attribute\validate.xlsx'
    wait = WebDriverWait(driver,30,poll_frequency=1,ignored_exceptions=[ElementNotVisibleException, ElementNotSelectableException, NoSuchElementException, StaleElementReferenceException])
    driver.get(base_url)
    wait.until(EC.element_to_be_clickable((By.ID, "oktaEmailInput"))).send_keys(demo_account['Account'])
    wait.until(EC.element_to_be_clickable((By.ID, "oktaSignInBtn"))).click()
    wait.until(EC.element_to_be_clickable((By.ID, "password-sign-in"))).send_keys(password)
    wait.until(EC.element_to_be_clickable((By.ID, 'onepass-submit-btn'))).click()
    time.sleep(5)
    # driver.get(driver.current_url)
 
    # try:
    #     authenticate_with_certificate('BOT DEC-D001a')
    # except Exception as e:
    #     print(str(e))
    #     try:
    #         authenticate_with_certificate('BOT DEC-D001a')
    #     except Exception as e:
    #         print(str(e)) 
    #         try:
    #             print('3rd time')
    #             authenticate_with_certificate('BOT DEC-D001a')
    #             print('worked')
    #         except Exception as e:
    #             print(str(e))
    #             keyboard = Controller()
    #             keyboard.press(Key.enter)
    #             keyboard.release(Key.enter)
            
    

# def submit_form():     
#     # timeout = 20 #wait time to load windows before quitting the program
#     # load_start_time = time.time()
#     # window = gw.getWindowsWithTitle('Submit Form - Google ')
#     # while not window and (time.time() - load_start_time) < timeout:
#     #     pyautogui.PAUSE = 0.5
#     #     window = gw.getWindowsWithTitle('Submit Form - Google ')

#     # if not window:
#     #     print('\nCHROME WINDOW DID NOT LOAD\n')
#     #     quit()

#     # # window[-1].activate()
#     # time.sleep(10)
#     authenticate_with_certificate('BOT DEC-D001a')

# def login_with_threading(demo_account):
#     t1_login = threading.Thread(target=login, args=(demo_account,))
#     t2_form = threading.Thread(target=submit_form)
#     print('line no. 65')
#     t1_login.start()
#     time.sleep(20)
#     print('line no.68')
#     t2_form.start()
#     time.sleep(20)
#     t1_login.join()
#     t2_form.join()

    

def email(excel_file_path):
    # Check and close Outlook if it's open
    try:
        # Kill outlook
        os.system('taskkill /im outlook.exe /f')
    except Exception as e:
            print(str(e))
            
    flag = False
    while not flag:
        try:
            outlook = win32.Dispatch('Outlook.application')
            mail = outlook.CreateItem(0)
            flag = True
        except Exception as e:
            print(str(e))

    # mail = outlook.CreateItem(0)
    mail.To = 'biki.dey@hpe.com;somaiah.kodimaniyanda-lava@hpe.com;srividya-d@hpe.com;ragul.subramani@hpe.com;kalaivanan.a@hpe.com;peng.wang3@hpe.com;jiaojiao.ding@hpe.com;jingz@hpe.com;weiwei.shao@hpe.com;mrunal.vinod@hpe.com'
    # mail.To = 'biki.dey@hpe.com'
    # ;somaiah.kodimaniyanda-lava@hpe.com'
    # mail.To = 'pranav-m.bhat@hpe.com'
    mail.Subject ="Demo account validation result " + str(datetime.now().strftime("%b %d"))
    mail.Body = "Dear Team, Please check the validation result as attached  Regards"
    mail.Attachments.Add(excel_file_path)
    mail.Send() 
    return 

# Define codeCompare function
def codeCompare(arr1, arr2):
    missingValue = ''
    for val in arr1:
        if val not in arr2:
            print("exist in default but missing in DXP:", val)
            missingValue += val + ','
    return missingValue[:-1]

# Define setDemoValue function
def setDemoValue(demo_account, BRDXP, GeoDXP, AttributeDXP, UserRightDXP):
    print('I am here')
    tempDemoAccounts = demo_account
    if BRDXP:
        tempDemoAccounts['BRDXP'] = BRDXP
    if GeoDXP:
        tempDemoAccounts['GeoDXP'] = GeoDXP
    if AttributeDXP:
        tempDemoAccounts['AttributeDXP'] = AttributeDXP
    if UserRightDXP:
        tempDemoAccounts['UserRightDXP'] = UserRightDXP
    demo_account = tempDemoAccounts
    print('last line')


def process_data(demo_account):
    
    
    current_url = driver.current_url
    if current_url == 'https://partner.hpe.com/group/prp/home':
        driver.get(demo_account['Simulation'])

    tempBR = None
    tempAttrArray = None
    tempURArray = None
    tempGeo = None
    tempBRDXP = None  # Initialize tempBRDXP
    tempGeoDXP = None  
    tempDXPAttrArray = None
    tempDXPURArray = None

    # Loop through elements with class 'aui-field-label'
    for l in range(len(driver.find_elements(By.CLASS_NAME, 'aui-field-label'))):
        label = driver.find_elements(By.CLASS_NAME, 'aui-field-label')[l].text
        if label == "BR":
            tempBR = driver.find_elements(By.CLASS_NAME, 'aui-helper-clearfix')[l]
            print('tempBR :', tempBR.text)
        elif label == "User EPI":
            tempURArray = driver.find_elements(By.CLASS_NAME, 'aui-helper-clearfix')[l]
            print('tempURArray :', tempURArray.text)
        elif label == "Extended Attributes":
            tempAttrArray = driver.find_elements(By.CLASS_NAME, 'aui-helper-clearfix')[l]
            print('tempAttrArray :', tempAttrArray.text)
        elif label == "Geography":
            tempGeo = driver.find_elements(By.CLASS_NAME, 'aui-helper-clearfix')[l]
            print('tempGeo :', tempGeo.text)

    DXPAttrArray = []  # Attr in DXP
    if tempAttrArray:
        for j in range(len(tempAttrArray.find_elements(By.TAG_NAME, 'li'))):
            DXPAttrArray.append(tempAttrArray.find_elements(By.TAG_NAME, 'li')[j].text)
        
        # print('DXPAttrArray :',type(DXPAttrArray))
        # print('Attribute_column :',demo_account['Attribute'])
        # print('DXPAttrArray :',DXPAttrArray)
        demo_account['Attribute'] = [item.strip() for item in demo_account['Attribute']]
        DXPAttrArray = [item.strip() for item in DXPAttrArray]
        tempDXPAttrArray = codeCompare(demo_account['Attribute'], DXPAttrArray)


    DXPURArray = []  # Initialize array to store User Rights in DXP
    if tempURArray:
        for k in range(len(tempURArray.find_elements(By.TAG_NAME, 'li'))):
            DXPURArray.append(tempURArray.find_elements(By.TAG_NAME, 'li')[k].text)
            print(type(tempURArray.find_elements(By.TAG_NAME, 'li')[k].text))
            
        # print('UserRight_column :',demo_account['UserRight'])
        # print('DXPURArray :',DXPURArray)
        demo_account['UserRight'] = [item.strip() for item in demo_account['UserRight']]
        DXPURArray = [item.strip() for item in DXPURArray]
        tempDXPURArray = codeCompare(demo_account['UserRight'], DXPURArray)
        # print(tempDXPURArray,'>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

    if tempBR:
        tempBRDXP = tempBR.find_elements(By.CLASS_NAME, 'aui-textboxlistentry-text')[0].text
        print('tempBRDXP :',tempBRDXP)
        if tempBRDXP == demo_account['BR']:
            tempBRDXP = ''

    if tempGeo:
        tempGeoDXP = tempGeo.find_elements(By.CLASS_NAME, 'aui-textboxlistentry-text')[0].text
        if tempGeoDXP == demo_account['Geo']:
            tempGeoDXP = ''

    # Perform setDemoValue function
    setDemoValue(demo_account, tempBRDXP, tempGeoDXP, tempDXPAttrArray, tempDXPURArray)

def tearDown():
    driver.close()

def generate_excel_report(demo_account,excel_file_path,sheet_name):
# Convert the dictionary to a DataFrame
    df = pd.DataFrame.from_dict(demo_account, orient='index').transpose()
    
    # Write the DataFrame to an Excel file
    df_old = pd.read_excel(excel_file_path,sheet_name=sheet_name)
    df = df_old.append(df)
    with pd.ExcelWriter(excel_file_path,engine='openpyxl',mode='a',if_sheet_exists='replace') as writer:
        df.to_excel(writer,sheet_name =sheet_name,index=False)

def tearDown():
    driver.close()

def remove_data_from_excel(file_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Iterate through each sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Clear data starting from the second row to keep the column names
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None

    # Save the modified Excel file
    wb.save(file_path)



# login()   
# process_data()
# generate_excel_report()
# email()

    

